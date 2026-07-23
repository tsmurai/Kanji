import openpyxl
from pathlib import Path

root = Path('c:/Users/つねひこ/Desktop/AI Kanji')
wb = openpyxl.load_workbook(root / '漢字の要.xlsx', data_only=True)
print('sheets:', wb.sheetnames)
for sheet_name in ['P6-', 'P48-50,68-93']:
    ws = wb[sheet_name]
    print('---', sheet_name, '---')
    print('max_row', ws.max_row, 'max_column', ws.max_column)
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 8), values_only=True):
        print(row)
